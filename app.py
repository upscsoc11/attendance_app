from reportlab.lib.pagesizes import A4
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
)
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

from flask import Flask, render_template, session, redirect, request, jsonify, session, send_from_directory, send_file, flash
from db.models import db, Staff, ExamDate, Attendance
from datetime import datetime
from auth import auth_bp, login_required, admin_required
from io import BytesIO
from openpyxl import Workbook
import os



app = Flask(__name__)

app.secret_key = os.getenv("SECRET_KEY", "dev-secret")

app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///attendance.db"

app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db.init_app(app)
app.register_blueprint(auth_bp)


@app.route("/")
def home():
    return redirect("/login")


@app.route("/certificate")
@login_required
def certificate():
    staff = Staff.query.order_by(Staff.name).all()
    dates = ExamDate.query.filter_by(active=True).order_by(ExamDate.exam_date).all()

    exam_name = dates[0].exam_name if dates else "NO ACTIVE EXAM"

    exam = dates[0] if dates else None

    return render_template(
    	"certificate.html",
    	exam_name=exam.exam_name if exam else "NO ACTIVE EXAM",
    	duty_type=exam.duty_type if exam else "N/A",
    	dates=dates,
    	staff=staff
    )


@app.route("/submit", methods=["POST"])
def submit_attendance():

    if "username" not in session:
        return jsonify(success=False, message="Unauthorized")

    username = session["username"]
    data = request.json
    persons = data.get("persons", [])

    exam = ExamDate.query.filter_by(active=True).first()

    # ❌ Prevent multiple submissions per exam per user

    duplicate_info = []

    for p in persons:
        id_no = p["id"]
        name = p["text"].split(" (")[0]
        new_dates = sorted(p["dates"])

        # Fetch all previous records for same person + same exam
        existing_records = Attendance.query.filter_by(
            id_no=id_no,
            exam_name=exam.exam_name
        ).all()

        existing_dates = []

        for record in existing_records:
            existing_dates.extend(record.exam_dates.split(", "))

        # Check overlap
        duplicate_dates = [d for d in new_dates if d in existing_dates]

        if duplicate_dates:
            duplicate_info.append(
                f"{name} ({', '.join(duplicate_dates)})"
            )

    if duplicate_info:
        return jsonify(
            success=False,
            message="Duplicate date(s) already submitted for: " + "; ".join(duplicate_info)
        )



    # 📁 PDF folder
    pdf_dir = "generated_pdfs"
    os.makedirs(pdf_dir, exist_ok=True)

    pdf_name = f"{username}_{exam.exam_name}_{datetime.now().strftime('%Y%m%d%H%M')}.pdf"
    pdf_path = os.path.join(pdf_dir, pdf_name)

    # 🧾 Save attendance rows
    for p in persons:
        att = Attendance(
            username=username,
            id_no=p["id"],
            name=p["text"].split(" (")[0],
            designation=p["text"].split("(")[-1].replace(")", ""),
            exam_name=exam.exam_name,
            exam_dates=", ".join(sorted(p["dates"])),
            pdf_path=pdf_path
        )
        db.session.add(att)

    db.session.commit()

    # 📄 PDF generation placeholder (next step)
    generate_pdf(pdf_path, persons, exam)

    return jsonify(success=True, pdf_url="/" + pdf_path)

@app.route("/generated_pdfs/<path:filename>")
def serve_pdf(filename):
    return send_from_directory("generated_pdfs", filename)

# =========================
# ADMIN DASHBOARD
# =========================

@app.route("/admin/dashboard")
@admin_required
def admin_dashboard():

    
    exam_filter = request.args.get("exam")

    if exam_filter:
        records = Attendance.query.filter_by(
            exam_name=exam_filter
        ).order_by(Attendance.created_at.desc()).all()
    else:
        records = Attendance.query.order_by(
            Attendance.created_at.desc()
        ).all()

    exams = db.session.query(Attendance.exam_name).distinct().all()
    exams = [e[0] for e in exams]

    return render_template(
        "admin_dashboard.html",
        records=records,
        exams=exams,
        selected_exam=exam_filter
    )


# =========================
# DELETE RECORD (ADMIN)
# =========================

@app.route("/admin/delete/<int:record_id>")
@admin_required
def delete_record(record_id):

    
    record = Attendance.query.get_or_404(record_id)

    db.session.delete(record)
    db.session.commit()

    return redirect("/admin/dashboard")

# =========================
# EXPORT TO EXCEL (ADMIN)
# =========================

@app.route("/admin/export")
@admin_required
def export_excel():

    
    exam_filter = request.args.get("exam")

    if exam_filter:
        records = Attendance.query.filter_by(
            exam_name=exam_filter
        ).order_by(Attendance.created_at.desc()).all()
    else:
        records = Attendance.query.order_by(
            Attendance.created_at.desc()
        ).all()

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Header row
    headers = [
        "ID No",
        "Name",
        "Designation",
        "Exam Name",
        "Exam Dates",
        "Submitted By",
        "Created At"
    ]
    ws.append(headers)

    # Data rows
    for r in records:
        ws.append([
            r.id_no,
            r.name,
            r.designation,
            r.exam_name,
            r.exam_dates,
            r.username,
            r.created_at.strftime("%Y-%m-%d %H:%M")
        ])

    # Save to memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = "attendance_report.xlsx"
    if exam_filter:
        filename = f"{exam_filter}_attendance.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
# =========================
# STAFF MANAGEMENT (ADMIN)
# =========================

@app.route("/admin/staff")
@login_required
def manage_staff():

    if session.get("role") != "admin":
        return redirect("/certificate")

    search = request.args.get("search")

    if search:
        staff_list = Staff.query.filter(
            Staff.name.ilike(f"%{search}%")
        ).order_by(Staff.id_no.asc()).all()
    else:
        staff_list = Staff.query.order_by(Staff.id_no.asc()).all()

    return render_template(
        "admin_staff.html",
        staff=staff_list,
        search=search
    )


@app.route("/admin/staff/add", methods=["POST"])
@login_required
def add_staff():

    if session.get("role") != "admin":
        return redirect("/certificate")

    id_no = request.form["id_no"]
    name = request.form["name"]
    designation = request.form["designation"]
    mts_no = request.form.get("mts_no")
    section = request.form.get("section")

    existing = Staff.query.filter_by(id_no=id_no).first()
    if existing:
        return redirect("/admin/staff")

    new_staff = Staff(
        id_no=id_no,
        name=name,
        designation=designation,
        mts_no=mts_no,
        section=section
    )

    db.session.add(new_staff)
    db.session.commit()

    return redirect("/admin/staff")


@app.route("/admin/staff/delete/<id_no>")
@login_required
def delete_staff(id_no):

    if session.get("role") != "admin":
        return redirect("/certificate")

    staff = Staff.query.get_or_404(id_no)

    db.session.delete(staff)
    db.session.commit()

    return redirect("/admin/staff")
    
# =========================
# BULK DELETE
# =========================

@app.route("/admin/staff/bulk_delete", methods=["POST"])
@login_required
def bulk_delete_staff():

    if session.get("role") != "admin":
        return redirect("/certificate")

    selected_ids = request.form.getlist("staff_ids")

    if not selected_ids:
        flash("No staff selected", "danger")
        return redirect("/admin/staff")

    for staff_id in selected_ids:
        staff = Staff.query.get(staff_id)
        if staff:
            db.session.delete(staff)

    db.session.commit()

    flash(f"Deleted {len(selected_ids)} staff record(s)", "success")

    return redirect("/admin/staff")    
    
    

# =========================
# IMPORT STAFF FROM EXCEL
# =========================

@app.route("/admin/staff/import", methods=["POST"])
@login_required
def import_staff():

    if session.get("role") != "admin":
        return redirect("/certificate")

    file = request.files.get("file")

    if not file:
        return redirect("/admin/staff")

    from openpyxl import load_workbook

    wb = load_workbook(file)
    ws = wb.active

    imported = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        id_no, name, designation, mts_no, section = row

        if not id_no or not name or not designation:
            skipped += 1
            continue

        existing = Staff.query.filter_by(id_no=str(id_no)).first()

        if existing:
            skipped += 1
            continue

        staff = Staff(
            id_no=str(id_no),
            name=str(name),
            designation=str(designation),
            mts_no=str(mts_no) if mts_no else None,
            section=str(section) if section else None
        )

        db.session.add(staff)
        imported += 1

    db.session.commit()

    return redirect("/admin/staff")



def generate_pdf(pdf_path, persons, exam):
    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=A4,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )

    styles = getSampleStyleSheet()
    story = []

    # Heading
    story.append(Paragraph("<b>C-XI SECTION</b>", styles["Normal"]))
    story.append(Spacer(1, 12))

    # Reference paragraph
    ref_text = (
        "Reference C-XI Section’s requisition for attending office in connection "
        f"with the <b>{exam.duty_type}</b> on "
        f"<b>{exam.exam_name}</b> Examination."
    )
    story.append(Paragraph(ref_text, styles["Normal"]))
    story.append(Spacer(1, 12))

    cert_text = (
        "It is certified that the undermentioned official(s) has/have actually "
        "attended office for performing the assigned work on the following date(s):"
    )
    story.append(Paragraph(cert_text, styles["Normal"]))
    story.append(Spacer(1, 12))

    # Table data
    table_data = [
        ["Sl. No.", "Name & Designation of the Officer", "Date(s)"]
    ]

    for i, p in enumerate(persons, start=1):
        table_data.append([
            str(i),
            p["text"],
            ", ".join(p["dates"])
        ])

    table = Table(table_data, colWidths=[50, 300, 120])
    table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))

    story.append(table)
    story.append(Spacer(1, 60))

    # Signature
    story.append(
        Paragraph(
            "(Signature of DS/US/DD/SO/PA/DPA/Consultant)<br/>C-XI Section",
            styles["Normal"]
        )
    )

    doc.build(story)


if __name__ == "__main__":
    app.run()
    #app.run(debug=True)
